import openpyxl


def load_data(filename):
    """从 Excel 文件读取销售数据"""
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        record["小计"] = record["数量"] * record["单价"]
        records.append(record)

    return records


def analyze_total(records):
    """计算总销售额"""
    return sum(r["小计"] for r in records)


def analyze_by_category(records):
    """按类别统计销售额"""
    categories = {}
    for r in records:
        cat = r["类别"]
        if cat not in categories:
            categories[cat] = 0
        categories[cat] += r["小计"]
    return categories


def analyze_by_product(records):
    """按商品统计销售数量和销售额"""
    products = {}
    for r in records:
        name = r["商品"]
        if name not in products:
            products[name] = {"数量": 0, "销售额": 0}
        products[name]["数量"] += r["数量"]
        products[name]["销售额"] += r["小计"]
    return products


def print_report(records):
    """打印完整的分析报告"""
    total = analyze_total(records)
    categories = analyze_by_category(records)
    products = analyze_by_product(records)

    print("=" * 50)
    print("        一月份销售数据分析报告")
    print("=" * 50)

    print(f"\n总销售额：{total:,.0f} 元")
    print(f"总记录数：{len(records)} 条")

    print("\n--- 按类别统计 ---")
    for cat, amount in categories.items():
        ratio = amount / total * 100
        print(f"  {cat}：{amount:>10,.0f} 元（占比 {ratio:.1f}%）")

    print("\n--- 按商品统计 ---")
    sorted_products = sorted(
        products.items(),
        key=lambda x: x[1]["销售额"],
        reverse=True,
    )
    for name, info in sorted_products:
        print(f"  {name}：卖出 {info['数量']} 件，销售额 {info['销售额']:,.0f} 元")

    top = sorted_products[0]
    print(f"\n最畅销商品：{top[0]}（销售额 {top[1]['销售额']:,.0f} 元）")


def save_report(products, filename):
    """把商品统计结果保存为 Excel 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "商品汇总"

    ws.append(["商品", "销售数量", "销售额"])

    sorted_products = sorted(
        products.items(),
        key=lambda x: x[1]["销售额"],
        reverse=True,
    )
    for name, info in sorted_products:
        ws.append([name, info["数量"], info["销售额"]])

    wb.save(filename)


def main():
    records = load_data("sales_data.xlsx")
    print_report(records)

    products = analyze_by_product(records)
    save_report(products, "product_summary.xlsx")
    print("\n商品汇总已保存到 product_summary.xlsx，可以用 Excel 打开查看")


main()
